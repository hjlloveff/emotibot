#!/usr/bin/python
#coding:utf-8
import requests
import json
from openpyxl import load_workbook
import sys, getopt
import time

#shadow url
dalurl = sys.argv[3]
#aws url
#dalurl = 'http://52.80.146.132:8885/dal'
appid = sys.argv[1]
payload = {}
payload['appid'] = appid
entities = []
payload.update(
    {
        'op':'upsert',
        'category': 'lq',
        'userRecordId': 0,
        'data': {
            'subop': 'defaultSubop'
        }
    }
)

def insertLq(start, end) :
#    entities.clear()
    entities = []
    for x in range(start, end):
        print '导入第', x,'条语料'
        corpus = ws.cell(row=x, column=1).value
        question = ws.cell(row=x, column=2).value
        item = {}
        item['content'] = corpus
        item['istemplate'] = True
        item['parentContent'] = question
        entities.append(item)
    payload['data']['entities'] = entities
    r = requests.post(dalurl, json=payload)
    if r.status_code == 200:
        result = json.loads(r.text)
    #    tmp = result['actualResults']

filename = sys.argv[2]
wb = load_workbook(filename)
ws = wb['Sheet1']

batchnum = 2000
rownum = ws.max_row
number = rownum
batch = rownum / batchnum
batchint = int(batch)
if batch > 1 :
    insertLq(2, batchnum)
    for i in range(batchint) :
        start = batchnum * (i+1)
        end = batchnum * (i+2)
        if end > rownum :
            end = rownum + 1
        insertLq(start, end)
else :
    insertLq(2, rownum + 1)
print '导入模板语料完成'
