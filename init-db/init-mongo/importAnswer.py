#!/usr/bin/python
#coding:utf-8
import requests
import json
from openpyxl import Workbook
from openpyxl import load_workbook
import sys, getopt
import time

#shadow url
dalurl = sys.argv[3]
#aws url
#dalurl = 'http://52.80.146.132:8885/dal'
appid = sys.argv[1]
deletepayload = {}
deletepayload.update(
    {
	    "op": "delete",
	    "category": "lq",
	    "appid": appid,
	    "userRecordId": 0,
	    "data": {
		    "subop": "conditionsSubop",
		    "conditions": {}
	    }
    }
)
# 删除模板appid下的所有扩写，回答，问题
r = requests.post(dalurl, json=deletepayload)
deletepayload['category'] = 'answer'
requests.post(dalurl, json=deletepayload)
deletepayload['category'] = 'sq'
requests.post(dalurl, json=deletepayload)

payload = {}
payload['appid'] = appid
entities = []
payload.update(
    {
        'op':'insert',
        'category': 'sq',
        'userRecordId': 0,
        'data': {
            'subop': 'defaultSubop'
        }
    }
)
answerPayload = {}
answerPayload.update(
    {
        'appid': appid,
        'op':'insert',
        'category':'answer',
        'userRecordId':0,
        'data': {
            'subop': 'defaultSubop'
        }
    }
)
answerentities = []
conditions = {}
data = {}

filename = sys.argv[2]
wb = load_workbook(filename)
ws = wb['Sheet1']

rownum = ws.max_row
questionmap = {}

def insertAnswer(answer, recordId) :
    answeritem['content'] = answer
    answeritem['parentRecordId'] = recordId
    answeritem['istemplate'] = True
  #  answerentities.clear()
    answerentities = []
    answerentities.append(answeritem)
    answerPayload['data']['entities'] = answerentities
    response = requests.post(dalurl, json=answerPayload)
    if response.status_code == 200 :
        answerresult = json.loads(response.text)
        answertmp = answerresult['actualResults']




for x in range(2, rownum + 1) :
    question = ws.cell(row=x, column=1).value
    answer = ws.cell(row=x, column=2).value
    item = {}
    item['content'] = question
    item['answers'] = [answer]
    item['istemplate'] = True
 #   entities.clear()
    entities = []
    entities.append(item)
    payload['data']['entities'] = entities

    if question in questionmap :

        insertAnswer(answer, questionmap[question])
    else :

        r = requests.post(dalurl, json=payload)
        if r.status_code == 200 :

            result = json.loads(r.text)
	    print result	
            if result[u'errno'] != 'OK' :
                continue
            tmp = result['actualResults']
            if tmp is not None:
                for y in tmp:
                    questionmap[question] = y
                    answeritem = {}
                    insertAnswer(answer, y)
                    print "insertAnswer"


#print('导入标准问题模板完成')
print '导入标准问题模板完成'
 
        
    
    
    

